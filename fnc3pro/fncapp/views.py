# views.py

import http.client
import json
import os, io, logging
import openpyxl
import uuid

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from fncapp.form import *
from django.urls import reverse_lazy, reverse
from django.db import transaction
from django.views.generic import CreateView, UpdateView, DeleteView, View
from utility.datatable import *
from django.contrib.auth.mixins import PermissionRequiredMixin
# Import dei tuoi model
from .models import *
from .services_openapi import connection_openapi, normalize_ateco

# =======================
#    INDEX
# =======================
def index(request):
    return render(request, 'index.html')

# =======================
#    AZIENDE
# =======================
def lista_aziende(request):
    # Conta il totale delle aziende
    aziende = Azienda.objects.all()
    totale_aziende = aziende.count()
    return render(request, 'aziende/lista_aziende.html', {'aziende': aziende, 'totale_aziende': totale_aziende})

logger = logging.getLogger(__name__)
def aziende_details(request, id_azienda):
    azienda = get_object_or_404(Azienda, pk=id_azienda)
    
    # Dipendenti associati all'azienda
    dipendenti = azienda.dipendenti.all()
    
    # Verifica prerequisiti (esempio di logica)
    # 1. Almeno un dipendente
    dipendenti_presenti = dipendenti.exists()

    # 2. Almeno un timesheet esistente per i dipendenti di questa azienda
    timesheets_azienda = Timesheet.objects.filter(dipendente__in=dipendenti)
    piani_presenti = timesheets_azienda.exists()

    # 3. Tutti i timesheet hanno ore_effettuate > 0
    ore_registrate = all(
        (ts.ore_effettuate is not None and ts.ore_effettuate > 0)
        for ts in timesheets_azienda
    ) if piani_presenti else False

    # Calcola i prerequisiti generali
    prerequisiti = dipendenti_presenti and piani_presenti and ore_registrate

    # Log utile per debug
    logger.info(f"Azienda: {azienda.nome}, Dipendenti: {dipendenti.count()}")
    # Passa i dati al template
    context = {
        'azienda': azienda,
        'dipendenti': dipendenti,
        'prerequisiti': prerequisiti,
    }
    return render(request, 'aziende/aziende_details.html', context)


def azienda_crud(request, id_azienda=None):
    """
    Creazione o modifica di un'azienda.
    """
    if id_azienda:
        azienda = get_object_or_404(Azienda, pk=id_azienda)
        logger.debug(f"Modifica Azienda ID: {id_azienda}")
    else:
        azienda = None
        logger.debug("Creazione di una nuova Azienda")

    if request.method == 'POST':
        nome = request.POST.get('azienda')
        partitaiva = request.POST.get('partitaiva')
        codice_ateco_id = request.POST.get('codiceateco')
        email = request.POST.get('email')
        telefono = request.POST.get('telefono')
        cellulare = request.POST.get('cellulare')
        indirizzosl = request.POST.get('indirizzosl')
        note = request.POST.get('note')

        logger.debug(f"POST data: nome={nome}, partitaiva={partitaiva}, codice_ateco_id={codice_ateco_id}, email={email}, telefono={telefono}, cellulare={cellulare}, indirizzosl={indirizzosl}, note={note}")

        # Recupera (o None) il CodiceAteco
        codice_ateco_obj = None
        if codice_ateco_id:
            try:
                codice_ateco_id = int(float(codice_ateco_id))  # Converte in intero dopo aver gestito eventuali decimali
                codice_ateco_obj = CodiceAteco.objects.get(pk=codice_ateco_id)
                logger.debug(f"Codice Ateco trovato: {codice_ateco_obj.codice}")
            except (ValueError, CodiceAteco.DoesNotExist):
                logger.error(f"Codice Ateco non valido: {codice_ateco_id}")
                codice_ateco_obj = None

        if azienda:
            # Aggiornamento
            logger.debug(f"Aggiornamento Azienda ID: {azienda.id}")
            azienda.nome = nome
            azienda.partitaiva = partitaiva
            azienda.codiceateco = codice_ateco_obj
            azienda.email = email
            azienda.telefono = telefono
            azienda.cellulare = cellulare
            azienda.indirizzosl = indirizzosl
            azienda.note = note
            azienda.save()
            logger.info(f"Azienda ID {azienda.id} aggiornata con successo")
        else:
            # Creazione
            logger.debug("Creazione di una nuova Azienda")
            azienda = Azienda.objects.create(
                nome=nome,
                partitaiva=partitaiva,
                codiceateco=codice_ateco_obj,
                email=email,
                telefono=telefono,
                cellulare=cellulare,
                indirizzosl=indirizzosl,
                note=note
            )
            logger.info(f"Azienda creata con ID {azienda.id}")

        return redirect('lista_aziende')

    # Se GET
    codici_ateco = CodiceAteco.objects.all()
    logger.debug(f"Rendering form con codici Ateco: {codici_ateco.count()} trovati")
    return render(request, 'aziende/azienda_crud.html', {
        'azienda': azienda,
        'codici_ateco': codici_ateco,
    })
    
def azienda_dipendenti(request, id_azienda):
    azienda = get_object_or_404(Azienda, pk=id_azienda)
    dipendenti = azienda.dipendenti.all()
    totale_dipendenti = dipendenti.count()  # Calcola il totale dei dipendenti
    context = {
        'azienda': azienda,
        'dipendenti': dipendenti,
        'totale_dipendenti': totale_dipendenti,  # Passa il totale al template
    }
    return render(request, 'aziende/azienda_dipendenti.html', context)

def elimina_azienda(request, id_azienda):
    """
    Elimina un'azienda con l'ID specificato.
    """
    azienda = get_object_or_404(Azienda, pk=id_azienda)
    
    # Eliminazione dell'azienda
    azienda.delete()
    
    # Messaggio di successo
    messages.success(request, f"L'azienda '{azienda.nome}' è stata eliminata con successo.")
    
    # Reindirizza alla lista aziende
    return redirect('lista_aziende') 

# =======================
#   DIPENDENTI
# =======================
def lista_dipendenti(request):
    dipendenti = Dipendente.objects.all()
    return render(request, 'dipendenti/lista_dipendenti.html', {'dipendenti': dipendenti})

def dipendenti_details(request, dipendente_id):
    dipendente = get_object_or_404(Dipendente, pk=dipendente_id)
    azienda = dipendente.azienda  # la ForeignKey ci consente di recuperare l'azienda direttamente
    return render(request, 'dipendenti/dipendenti_details.html', {
        'dipendente': dipendente,
        'azienda': azienda
    })

def dipendenti_crud(request, dipendente_id=None):
    if dipendente_id:
        dipendente = get_object_or_404(Dipendente, pk=dipendente_id)
    else:
        dipendente = None

    if request.method == "POST":
        nome = request.POST.get('nome')
        cognome = request.POST.get('cognome')
        sesso = request.POST.get('sesso')
        datanascita = request.POST.get('datanascita') or None
        luogonascita = request.POST.get('luogonascita')
        provincianascita = request.POST.get('provincianascita')
        codicefiscale = request.POST.get('codicefiscale')
        dataassunzione = request.POST.get('dataassunzione') or None
        datacessazione = request.POST.get('datacessazione') or None
        tipocontratto = request.POST.get('tipocontratto')
        livellocontratto = request.POST.get('livellocontratto')
        oresettimanali = request.POST.get('oresettimanali') or None
        percparttime = request.POST.get('percparttime') or None
        qualifica = request.POST.get('qualifica')
        mansione = request.POST.get('mansione')
        note = request.POST.get('note')
        email = request.POST.get('email')
        telefono = request.POST.get('telefono')
        azienda_id = request.POST.get('idazienda')

        # Verifica che l'azienda sia valida
        azienda = None
        if azienda_id:
            try:
                azienda = get_object_or_404(Azienda, pk=azienda_id)
            except Exception as e:
                return JsonResponse({'error': True, 'msgResponse': f'Azienda non valida: {str(e)}'})

        if not azienda:
            return JsonResponse({'error': True, 'msgResponse': 'L\'azienda è obbligatoria per salvare il dipendente.'})

        if dipendente:
            # Aggiorno esistente
            dipendente.nome = nome
            dipendente.cognome = cognome
            dipendente.sesso = sesso
            dipendente.datanascita = datanascita
            dipendente.luogonascita = luogonascita
            dipendente.provincianascita = provincianascita
            dipendente.codicefiscale = codicefiscale
            dipendente.dataassunzione = dataassunzione
            dipendente.datacessazione = datacessazione
            dipendente.tipocontratto = tipocontratto
            dipendente.livellocontratto = livellocontratto
            dipendente.oresettimanali = oresettimanali
            dipendente.percparttime = percparttime
            dipendente.qualifica = qualifica
            dipendente.mansione = mansione
            dipendente.note = note
            dipendente.email = email
            dipendente.telefono = telefono
            dipendente.azienda = azienda
            dipendente.save()
        else:
            # Creo nuovo
            Dipendente.objects.create(
                nome=nome,
                cognome=cognome,
                sesso=sesso,
                datanascita=datanascita,
                luogonascita=luogonascita,
                provincianascita=provincianascita,
                codicefiscale=codicefiscale,
                dataassunzione=dataassunzione,
                datacessazione=datacessazione,
                tipocontratto=tipocontratto,
                livellocontratto=livellocontratto,
                oresettimanali=oresettimanali,
                percparttime=percparttime,
                qualifica=qualifica,
                mansione=mansione,
                note=note,
                email=email,
                telefono=telefono,
                azienda=azienda,
            )
        return redirect('lista_dipendenti')

    aziende = Azienda.objects.all()
    return render(request, 'dipendenti/dipendenti_crud.html', {
        'dipendente': dipendente,
        'aziende': aziende,
    })
    
def dipendenti_delete(request, dipendente_id):
    """
    View per la cancellazione di un dipendente.
    """
    try:
        dipendente = get_object_or_404(Dipendente, pk=dipendente_id)
        dipendente.delete()
        return JsonResponse({'error': False, 'msgResponse': 'Dipendente eliminato con successo.'})
    except Exception as e:
        return JsonResponse({'error': True, 'msgResponse': f'Errore durante l\'eliminazione del dipendente: {str(e)}'})


# =======================
#   PIANI FORMATIVI
# =======================
def lista_piani(request):
    piani = PianoFormativo.objects.all()
    return render(request, 'piani/lista_pianiformativi.html', {'piani': piani})

def piano_details(request, id_piano):
    piano = get_object_or_404(PianoFormativo, pk=id_piano)
    return render(request, 'piani/piano_details.html', {'piano': piano})

def piano_crud(request, id_piano=None):
    if id_piano:
        piano = get_object_or_404(PianoFormativo, pk=id_piano)
    else:
        piano = None

    if request.method == "POST":
        nome = request.POST.get('titolo')
        data_inizio = request.POST.get('data_inizio') or None
        data_fine = request.POST.get('data_fine') or None
        ore_totali = request.POST.get('ore_totali') or None
        descrizione = request.POST.get('descrizione')

        progetto_id = request.POST.get('progetto')  # Se vuoi scegliere il progetto da un select
        fondo_id = request.POST.get('fondo')

        # Recupera Progetto (opzionale)
        if progetto_id:
            progetto = get_object_or_404(Progetto, pk=progetto_id)
        else:
            progetto = None

        # Recupera Fondo (opzionale)
        if fondo_id:
            fondo = get_object_or_404(TipoFondo, pk=fondo_id)
        else:
            fondo = None

        if piano:
            # Aggiorno
            piano.nome = nome
            piano.data_inizio = data_inizio
            piano.data_fine = data_fine
            piano.ore_totali = ore_totali
            piano.descrizione = descrizione
            piano.progetto = progetto  # se None, DB avrà progetto_id = NULL
            piano.fondo = fondo
            piano.save()
        else:
            # Crea
            PianoFormativo.objects.create(
                nome=nome,
                data_inizio=data_inizio,
                data_fine=data_fine,
                ore_totali=ore_totali,
                descrizione=descrizione,
                progetto=progetto,   # None se non selezionato
                fondo=fondo,
            )
        return redirect('lista_pianiformativi')

    progetti = Progetto.objects.all()
    fondi = TipoFondo.objects.all()
    return render(request, 'piani/pianiformativi_crud.html', {
        'pianoformativo': piano,
        'progetti': progetti,
        'fondi': fondi,
    })

def piano_elimina(request, id_piano):
    piano = get_object_or_404(PianoFormativo, pk=id_piano)
    piano.delete()
    return redirect('lista_pianiformativi')


# =======================
#   PROGETTI
# =======================
def lista_progetti(request):
    """
    Mostra la lista dei progetti e i piani associati (se presenti).
    """
    progetti = Progetto.objects.all().prefetch_related('piani_formativi')
    
    # Creiamo una piccola struttura in cui ogni progetto avrà anche
    # la lista dei nomi dei piani formativi associati
    progetti_render = []
    for prj in progetti:
        piani_nomi = [pf.nome for pf in prj.piani_formativi.all()]
        progetti_render.append({
            'id_progetto': prj.id,
            'nome': prj.nome,
            'data_avvio': prj.data_avvio,
            'stato': prj.stato,
            'descrizione': prj.descrizione,
            'piani_nomi': piani_nomi
        })

    return render(request, 'progetti/lista_progetti.html', {
        'progetti': progetti_render
    })

def progetto_details(request, id_progetto):
    progetto = get_object_or_404(Progetto, pk=id_progetto)
    return render(request, 'progetti/progetto_details.html', {'progetto': progetto})

def progetto_crud(request, id_progetto=None):
    if id_progetto:
        progetto = get_object_or_404(Progetto, pk=id_progetto)
    else:
        progetto = None

    if request.method == "POST":
        nome = request.POST.get('nome')
        data_avvio = request.POST.get('data_avvio') or None
        stato = request.POST.get('stato')
        descrizione = request.POST.get('descrizione')

        if progetto:
            progetto.nome = nome
            progetto.data_avvio = data_avvio
            progetto.stato = stato
            progetto.descrizione = descrizione
            progetto.save()
        else:
            Progetto.objects.create(
                nome=nome,
                data_avvio=data_avvio,
                stato=stato,
                descrizione=descrizione,
            )
        return redirect('lista_progetti')

    return render(request, 'progetti/progetti_crud.html', {'progetto': progetto})

def progetto_elimina(request, id_progetto):
    progetto = get_object_or_404(Progetto, pk=id_progetto)
    progetto.delete()
    return redirect('lista_progetti')



# =======================
#   MODULI (CRUD)
# =======================

def lista_moduli(request):
    moduli = Modulo.objects.all()
    return render(request, 'moduli/lista_moduli.html', {'moduli': moduli})

def modulo_details(request, id_modulo):
    modulo = get_object_or_404(Modulo, pk=id_modulo)
    return render(request, 'moduli/modulo_details.html', {'modulo': modulo})

def modulo_crud(request, id_modulo=None):
    """
    Creazione o modifica di un modulo.
    """
    if id_modulo:
        modulo = get_object_or_404(Modulo, pk=id_modulo)
    else:
        modulo = None

    if request.method == 'POST':
        nome = request.POST.get('nome')
        descrizione = request.POST.get('descrizione')
        # altri campi, se previsti

        if modulo:
            # Aggiornamento
            modulo.nome = nome
            modulo.descrizione = descrizione
            modulo.save()
        else:
            # Creazione
            Modulo.objects.create(
                nome=nome,
                descrizione=descrizione,
            )
        return redirect('lista_moduli')

    return render(request, 'moduli/moduli_crud.html', {
        'modulo': modulo
    })

def modulo_elimina(request, id_modulo):
    modulo = get_object_or_404(Modulo, pk=id_modulo)
    modulo.delete()
    return redirect('lista_moduli')




# =======================
#   Wizard di associazione
#   (esempio con Progetti e Piani)
# =======================
def wizard_propiani(request):
    """
    Esempio di 'wizard' che associa più Piani Formativi ad un Progetto.
    """
    if request.method == "POST":
        progetto_id = request.POST.get('selectProgetto')
        piani_scelti = request.POST.getlist('pianiSelezionati')  # array di id dei PianoFormativo

        progetto = get_object_or_404(Progetto, pk=progetto_id)
        # Svuota tutti i piani attuali (se vuoi un'associazione "one-to-many" diretta).
        # Oppure, se vuoi solo aggiungere, non svuotare.
        PianoFormativo.objects.filter(progetto=progetto).update(progetto=None)

        # Assegna i piani scelti a questo progetto
        piani = PianoFormativo.objects.filter(pk__in=piani_scelti)
        for p in piani:
            p.progetto = progetto
            p.save()

        return redirect('lista_progetti')

    # GET
    progetti = Progetto.objects.all()
    piani = PianoFormativo.objects.all()
    return render(request, 'associazioni/wizard_propiani.html', {
        'progetti': progetti,
        'piani': piani,
    })


# =======================
#   Allegati / Export
# =======================

def pagina_generazione_allegato(request, id_azienda):
    azienda = get_object_or_404(Azienda, pk=id_azienda)
    dipendenti = azienda.dipendenti.all()

    timesheets_azienda = Timesheet.objects.filter(dipendente__in=dipendenti)
    ore_registrate = False
    if timesheets_azienda.exists():
        ore_registrate = all(
            (ts.ore_effettuate is not None and ts.ore_effettuate > 0)
            for ts in timesheets_azienda
        )

    # Calcoliamo i piani associati
    piani_set = set()
    for ts in timesheets_azienda.select_related('piano_modulo__piano_formativo'):
        if ts.piano_modulo and ts.piano_modulo.piano_formativo:
            piani_set.add(ts.piano_modulo.piano_formativo)
    piani = list(piani_set)

    # Leggiamo il moltiplicatore dal GET
    moltiplicatore = request.GET.get('moltiplicatore', '')

    context = {
        'azienda': azienda,
        'dipendenti': dipendenti,
        'piani': piani,
        'ore_lavorative_verificate': ore_registrate,
        'moltiplicatore': moltiplicatore,  # lo passiamo al template
    }
    return render(request, 'allegati/genera_allegato2bis.html', context)

def seleziona_azienda_allegato(request):
    """
    View per selezionare un'azienda e reindirizzare alla pagina di generazione allegato.
    """
    if request.method == 'POST':
        id_azienda = request.POST.get('id_azienda')
        if id_azienda:
            # Reindirizza alla view `pagina_generazione_allegato`
            return redirect('pagina_generazione_allegato', id_azienda=id_azienda)
        else:
            messages.error(request, "Seleziona un'azienda valida.")

    # Recupera tutte le aziende
    aziende = Azienda.objects.all()
    context = {'aziende': aziende}
    return render(request, 'aziende/seleziona_azienda_allegato.html', context)

def genera_allegato_excel(request, id_azienda):
    import os, io
    import openpyxl
    from django.http import HttpResponse
    from django.shortcuts import get_object_or_404
    from .models import Azienda, Timesheet

    # 1) Troviamo il percorso del template XLSX
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(base_dir, "staticfiles/documenti/Allegato_2_bis.xlsx")

    # 2) Carichiamo il file Excel
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active  # o wb["NomeFoglio"] se hai un foglio specifico

    # 3) Recuperiamo dati
    azienda = get_object_or_404(Azienda, pk=id_azienda)
    dipendenti = azienda.dipendenti.all()
    timesheets = Timesheet.objects.filter(dipendente__in=dipendenti).select_related("dipendente", "piano_modulo")

    # Lettura del moltiplicatore da querystring (se presente: ?moltiplicatore=0.6 ...)
    moltiplicatore_str = request.GET.get("moltiplicatore", "0.0")
    moltiplicatore = float(moltiplicatore_str) if moltiplicatore_str else 0.0

    # 4) Intestazione extra in A3 (se A3 è la top-left cell unita e scrivibile)
    ws["A3"] = "=COUNTA(A5:A5000)"

    # 5) Scriviamo i dati a partire dalla riga 5 (puoi modificare se serve)
    start_row = 5
    for i, ts in enumerate(timesheets, start=0):
        current_row = start_row + i

        d = ts.dipendente
        
        # Livello contrattuale dal dipendente
        livello_inq = d.livellocontratto or "N/D"
        # N° ore rimodulazione
        ore = float(ts.ore_effettuate or 0)
        # Quote orarie
        retrib_oraria = float(ts.importo_retributivo or 0)
        contrib_oraria = float(ts.importo_contributivo or 0)

        # Calcoli
        totale_retribuzione = ore * retrib_oraria  # (F)
        totale_contributo = ore * contrib_oraria   # (G)
        contributo_totale = totale_retribuzione + totale_contributo  # (H)

        # Compiliamo le colonne (A..H)
        ws.cell(row=current_row, column=1).value = d.codicefiscale       # A
        ws.cell(row=current_row, column=2).value = livello_inq           # B
        ws.cell(row=current_row, column=3).value = ore                   # C
        ws.cell(row=current_row, column=4).value = retrib_oraria         # D
        ws.cell(row=current_row, column=5).value = contrib_oraria        # E
        ws.cell(row=current_row, column=6).value = round(totale_retribuzione, 2)  # F
        ws.cell(row=current_row, column=7).value = round(totale_contributo, 2)     # G
        ws.cell(row=current_row, column=8).value = round(contributo_totale, 2)     # H

    # 6) Se vuoi scrivere la formula in F3 come SUM(F5:F5000)*moltiplicatore
    ws["F3"] = f"=SUM(F5:F5000)*{moltiplicatore}"

    # 7) Salviamo in memoria e inviamo come download
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Allegato_2_bis_{azienda.nome}.xlsx".replace(" ", "_")
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# =======================
#   CALCOLI DI CONTRIBUTO
# =======================
def calcola_contributo(codice_fiscale, ore_lavorate, quota_retribuzione_oraria, quota_contribuzione_oraria):
    """
    Esempio: calcola il contributo totale per un lavoratore
    (logica semplificata rispetto al mock originale).
    """
    ore_lavorate = ore_lavorate or 0
    quota_retribuzione_oraria = quota_retribuzione_oraria or 0
    quota_contribuzione_oraria = quota_contribuzione_oraria or 0

    totale_retribuzione = ore_lavorate * quota_retribuzione_oraria
    totale_contribuzione = ore_lavorate * quota_contribuzione_oraria
    contributo_totale = totale_retribuzione + totale_contribuzione
    return {
        "codice_fiscale": codice_fiscale,
        "totale_retribuzione": round(totale_retribuzione, 2),
        "totale_contribuzione": round(totale_contribuzione, 2),
        "contributo_totale": round(contributo_totale, 2),
    }

def genera_dati_contributo(request, id_azienda):
    """
    Esempio di pagina che mostra un riepilogo di calcolo contributi
    per i dipendenti di un'azienda in base ai Timesheet.
    """
    azienda = get_object_or_404(Azienda, pk=id_azienda)
    dipendenti = azienda.dipendenti.all()

    risultati = []
    for dipendente in dipendenti:
        timesheets = Timesheet.objects.filter(dipendente=dipendente)
        for ts in timesheets:
            ore_lavorate = ts.ore_effettuate or 0
            # Trattiamo 'importo_retributivo' e 'importo_contributivo' come tariffe orarie
            quota_retribuzione_oraria = ts.importo_retributivo or 0
            quota_contribuzione_oraria = ts.importo_contributivo or 0

            risultato = calcola_contributo(
                codice_fiscale=dipendente.codicefiscale,
                ore_lavorate=ore_lavorate,
                quota_retribuzione_oraria=quota_retribuzione_oraria,
                quota_contribuzione_oraria=quota_contribuzione_oraria,
            )
            risultati.append(risultato)

    return render(request, 'allegati/contributo_dati.html', {
        "azienda": azienda,
        "risultati": risultati,
    })


def wizard_allegato2bis(request):
    """
    Un'unica view per gestire:
    - GET: restituisce la pagina HTML con i 3 "step" (nascosti/visibili via JS).
    - POST (AJAX): riceve il file Excel, fa il parsing, restituisce i dati in JSON
      per mostrarli nello Step 3 (nuovi vs duplicati).
    - POST finale: quando l'utente conferma, crea/aggiorna i dipendenti su DB.
    """

    if request.method == 'POST':
        # Distinguere i 2 tipi di POST:
        # 1) POST "parse_file" (AJAX) -> analizza file Excel
        # 2) POST "confirm_import" -> l'utente ha confermato

        action = request.POST.get('action')

        if action == 'parse_file':
            # 1) Riceviamo i dati del form in AJAX
            azienda_id = request.POST.get('azienda_id')
            file_excel = request.FILES.get('file_excel')
            if not (azienda_id and file_excel):
                return JsonResponse({'error': 'Missing data'}, status=400)

            # Recupero azienda
            azienda = get_object_or_404(Azienda, pk=azienda_id)

            # Caricamento con openpyxl
            wb = openpyxl.load_workbook(file_excel)
            ws = wb.active

            nuovi = []
            duplicati = []

            start_row = 5
            row_index = start_row

            while True:
                cf = ws.cell(row=row_index, column=1).value  # CF
                if not cf:
                    break
                livello_inq = ws.cell(row=row_index, column=2).value
                ore_eff = ws.cell(row=row_index, column=3).value
                retrib = ws.cell(row=row_index, column=4).value
                contrib = ws.cell(row=row_index, column=5).value

                try:
                    dip = Dipendente.objects.get(codicefiscale=cf, azienda=azienda)
                    duplicati.append({
                        'dipendente_id': dip.id,
                        'codicefiscale': cf,
                        'livello_inq_attuale': dip.livellocontratto,
                        'livello_inq_nuovo': livello_inq,
                        'ore_effettuate': ore_eff or 0,
                        'retrib_oraria': retrib or 0,
                        'contrib_oraria': contrib or 0
                    })
                except Dipendente.DoesNotExist:
                    nuovi.append({
                        'codicefiscale': cf,
                        'livello_inq': livello_inq,
                        'ore_effettuate': ore_eff or 0,
                        'retrib_oraria': retrib or 0,
                        'contrib_oraria': contrib or 0
                    })

                row_index += 1

            # Torniamo i risultati in JSON
            return JsonResponse({
                'success': True,
                'nuovi': nuovi,
                'duplicati': duplicati,
            })

        elif action == 'confirm_import':
            # 2) L'utente ha cliccato "Conferma" nello step 3 e 
            #    ci manda i dati di "nuovi" e "duplicati" (in JSON o form).
            import json
            azienda_id = request.POST.get('azienda_id')
            if not azienda_id:
                return JsonResponse({'error': 'Nessuna azienda selezionata'}, status=400)

            azienda = get_object_or_404(Azienda, pk=azienda_id)

            nuovi_json = request.POST.get('nuovi')      # stringa JSON
            duplicati_json = request.POST.get('duplicati')  # stringa JSON
            duplicati_da_aggiornare = request.POST.getlist('duplicati_update[]', [])

            if not (nuovi_json and duplicati_json):
                return JsonResponse({'error': 'Mancano dati'}, status=400)

            nuovi = json.loads(nuovi_json)
            duplicati = json.loads(duplicati_json)

            # 2.a) Creazione dei nuovi dipendenti
            from .models import Dipendente, Timesheet
            for d in nuovi:
                Dipendente.objects.create(
                    codicefiscale=d['codicefiscale'],
                    nome='DaAllegato',
                    cognome='DaAllegato',
                    livellocontratto=d['livello_inq'] or '',
                    azienda=azienda,
                )
                # se vuoi creare timesheet, fallo qui

            # 2.b) Aggiornamento duplicati spuntati
            for dup in duplicati:
                if str(dup['dipendente_id']) in duplicati_da_aggiornare:
                    dip = Dipendente.objects.get(pk=dup['dipendente_id'])
                    dip.livellocontratto = dup['livello_inq_nuovo'] or dip.livellocontratto
                    dip.save()
                    # se vuoi timesheet:
                    # Timesheet.objects.create(...)

            return JsonResponse({'success': True})

    # Se GET, restituiamo l’unico template con i 3 step (in JS)
    from .models import Azienda
    aziende = Azienda.objects.all()
    return render(request, 'dipendenti/allegato2bis_wizard_unapagina.html', {
        'aziende': aziende
    })
    
    
    

def carica_azienda_openapi(request):
    partitaiva = request.GET.get('partitaiva', '').strip()
    if not partitaiva:
        return JsonResponse({'error': True, 'msgResponse': 'Partita IVA non fornita.'})

    # 1) Recupero dati da OpenAPI => tool="2"
    try:
        data = connection_openapi("2", partitaiva)
    except Exception as e:
        return JsonResponse({'error': True, 'msgResponse': f'Errore di connessione: {str(e)}'})

    # 2) Controlli sulla risposta
    if not isinstance(data, dict) or 'success' not in data:
        return JsonResponse({'error': True, 'msgResponse': 'Risposta inattesa da OpenAPI.'})

    if data['success'] == False:
        err_code = data.get('error', None)
        err_msg = data.get('message', '')
        if err_code == 207:
            return JsonResponse({'error': True, 'msgResponse': 'CF/PIVA non validi (207).'})
        elif err_code == 208:
            return JsonResponse({'error': True, 'msgResponse': 'Credito insufficiente (208).'})
        elif err_code == 209 or err_msg == 'internal login error':
            return JsonResponse({'error': True, 'msgResponse': 'Non sei autorizzato (209). Contatta area sviluppo.'})
        elif err_msg == 'no company found':
            return JsonResponse({'error': True, 'msgResponse': 'Azienda non trovata.'})
        else:
            return JsonResponse({'error': True, 'msgResponse': f'Errore generico: {err_msg}'})

    # Se success == True, estraiamo i dati
    info_data = data.get('data', {})
    if not info_data:
        return JsonResponse({'error': True, 'msgResponse': 'Nessun dato azienda ricevuto.'})

    denominazione = info_data.get('denominazione', '')
    piva = info_data.get('piva', '')
    cf = info_data.get('cf', '')
    indirizzo = info_data.get('indirizzo', '')
    comune = info_data.get('comune', '')
    provincia = info_data.get('provincia', '')
    cap = info_data.get('cap', '')
    stato_attivita = info_data.get('stato_attivita', '')

    dettaglio = info_data.get('dettaglio', {})
    codice_ateco_str = dettaglio.get('codice_ateco', '')
    pec = dettaglio.get('pec', '')

    # Normalizziamo il codice ATECO
    codice_ateco_norm = normalize_ateco(codice_ateco_str)
    
    # 3) Verifichiamo se esiste azienda con questa P.IVA
    azienda_esistente = Azienda.objects.filter(partitaiva=piva).first()
    if azienda_esistente:
        # Già registrata: calcoliamo le differenze
        differenze = {}
        print("DEBUG - Confronto nome DB='", azienda_esistente.nome, "' vs API='", denominazione, "'")
        if azienda_esistente.nome != denominazione:
            differenze['nome'] = {
                'db': azienda_esistente.nome or '',
                'api': denominazione or ''
            }
        if azienda_esistente.indirizzosl != indirizzo:
            differenze['indirizzosl'] = {
                'db': azienda_esistente.indirizzosl or '',
                'api': indirizzo or ''
            }
        if (azienda_esistente.email or '') != (pec or ''):
            differenze['email'] = {
                'db': azienda_esistente.email or '',
                'api': pec or ''
            }

        return JsonResponse({
            'error': False,
            'db': 'yes',
            'msgResponse': 'Azienda già registrata. Vuoi aggiornare i dati? (procedura a pagamento!)',
            'azienda': azienda_esistente.nome,
            'partitaiva': azienda_esistente.partitaiva,
            'differenze': differenze,
        })

    # 4) Creiamo nuova Azienda (se non esiste a DB)
    ateco_obj = CodiceAteco.objects.filter(codice=codice_ateco_norm).first()
    nuova = Azienda.objects.create(
        nome=denominazione,
        partitaiva=piva,
        codiceateco=ateco_obj,
        email=pec,
        indirizzosl=indirizzo,
        note=f"Comune: {comune}, Prov: {provincia}, CAP: {cap}, StatoAttività: {stato_attivita}."
    )

    return JsonResponse({
        'error': False,
        'db': 'no',
        'msgResponse': 'Azienda inserita con successo.',
        'azienda': denominazione,
        'partitaiva': piva,
        'codiceateco': codice_ateco_norm,
        'indirizzosl': indirizzo,
        'status_societa': stato_attivita,
    })


def update_azienda_openapi(request):
    partitaiva = request.GET.get('partitaiva', '').strip()
    if not partitaiva:
        return JsonResponse({'error': True, 'msgResponse': 'Parametri non corretti (p.iva mancante).'})

    # 1) Chiamata a OpenAPI => tool="2"
    try:
        data = connection_openapi("2", partitaiva)
    except Exception as e:
        return JsonResponse({'error': True, 'msgResponse': f'Errore di connessione: {str(e)}'})

    if not isinstance(data, dict) or 'success' not in data:
        return JsonResponse({'error': True, 'msgResponse': 'Risposta inattesa da OpenAPI.'})
    
    if data['success'] == False:
        err_code = data.get('error', None)
        err_msg = data.get('message', '')
        return JsonResponse({'error': True, 'msgResponse': f'Errore da API: {err_msg} (codice {err_code})'})

    info_data = data.get('data', {})
    if not info_data:
        return JsonResponse({'error': True, 'msgResponse': 'Dati azienda mancanti.'})

    denominazione = info_data.get('denominazione', '')
    piva = info_data.get('piva', '')
    cf = info_data.get('cf', '')
    indirizzo = info_data.get('indirizzo', '')
    comune = info_data.get('comune', '')
    provincia = info_data.get('provincia', '')
    cap = info_data.get('cap', '')
    stato_attivita = info_data.get('stato_attivita', '')

    dettaglio = info_data.get('dettaglio', {})
    codice_ateco_str = dettaglio.get('codice_ateco', '')
    pec = dettaglio.get('pec', '')
    codice_ateco_norm = normalize_ateco(codice_ateco_str)

    # 2) Cerchiamo l'azienda in DB
    azienda = Azienda.objects.filter(partitaiva=piva).first()
    if not azienda:
        return JsonResponse({'error': True, 'msgResponse': 'Azienda non presente in DB: impossibile aggiornarla.'})

    # 3) Aggiorniamo
    azienda.nome = denominazione
    if codice_ateco_norm:
        ateco_obj = CodiceAteco.objects.filter(codice=codice_ateco_norm).first()
        azienda.codiceateco = ateco_obj
    azienda.email = pec
    azienda.indirizzosl = indirizzo
    azienda.note = f"Comune: {comune}, Prov: {provincia}, CAP: {cap}, StatoAttivita: {stato_attivita}"
    azienda.save()

    return JsonResponse({
        'error': False,
        'msgResponse': 'Azienda aggiornata con successo.',
        'azienda': denominazione,
        'partitaiva': piva,
        'codiceateco': codice_ateco_norm,
        'indirizzosl': indirizzo,
        'status_societa': stato_attivita,
    })




#tonyboy 

class ProgettoCreateView(CreateView):
    model = Progetto
    form_class = ProgettoForm
    template_name = 'progetti/progetto_form.html'
    success_url = reverse_lazy('lista_progetti')
    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['piani_formativi'] = PianoFormativoInlineFormSet(self.request.POST, instance=self.object, prefix='piani_formativi')
        else:
            data['piani_formativi'] = PianoFormativoInlineFormSet(instance=self.object, prefix='piani_formativi')
        return data
    @transaction.atomic
    def form_valid(self, form):
        context = self.get_context_data()
        piani_formativi = context['piani_formativi']
        if not form.is_valid() or not piani_formativi.is_valid():
            return self.form_invalid(form)
        self.object = form.save()
        piani_formativi.instance = self.object
        piani_formativi.save()
        return redirect(self.success_url)
class ProgettoUpdateView(UpdateView):
    model = Progetto
    form_class = ProgettoForm
    template_name = 'progetti/progetto_form.html'
    success_url = reverse_lazy('lista_progetti')
    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['piani_formativi'] = PianoFormativoInlineFormSet(self.request.POST, instance=self.object, prefix='piani_formativi')
        else:
            data['piani_formativi'] = PianoFormativoInlineFormSet(instance=self.object, prefix='piani_formativi')
        return data
    @transaction.atomic
    def form_valid(self, form):
        context = self.get_context_data()
        piani_formativi = context['piani_formativi']
        if not form.is_valid() or not piani_formativi.is_valid():
            return self.form_invalid(form)
        self.object = form.save()
        piani_formativi.instance = self.object
        piani_formativi.save()
        return redirect(self.success_url)
def search_azienda(request):
    search_term = request.GET.get('term', '')
    page = request.GET.get('page', 1)
    page_size = 10
    queryset = Azienda.objects.filter(
        nome__icontains=search_term
    ).order_by('nome')
    total = queryset.count()
    start = (int(page) - 1) * page_size
    end = start + page_size
    results = [
        {
            'id': elem.pk,
            'text': elem.nome,
        }
        for elem in queryset[start:end]
    ]
    return JsonResponse({
        'results': results,
        'pagination': {'more': end < total}
    })
def search_modulo(request):
    search_term = request.GET.get('term', '')
    page = request.GET.get('page', 1)
    page_size = 10
    queryset = Modulo.objects.filter(
        nome__icontains=search_term
    ).order_by('nome')
    total = queryset.count()
    start = (int(page) - 1) * page_size
    end = start + page_size
    results = [
        {
            'id': elem.pk,
            'text': elem.nome,
        }
        for elem in queryset[start:end]
    ]
    return JsonResponse({
        'results': results,
        'pagination': {'more': end < total}
    })
    
# =======================
#    TIMESHEET
# =======================
def seleziona_azienda_timesheet(request):
    """
    Mostra una lista di aziende tra cui l'utente può scegliere.
    Una volta selezionata, reindirizza alla pagina corretta.
    """
    aziende = Azienda.objects.all()  # Recupera tutte le aziende
    if request.method == 'POST':
        # Recupera l'ID dell'azienda selezionata
        id_azienda = request.POST.get('id_azienda')
        if id_azienda:
            # Salva l'azienda in sessione (opzionale)
            request.session['azienda_selezionata'] = id_azienda
            return redirect('timesheet_per_azienda', id_azienda=id_azienda)
        else:
            messages.error(request, "Seleziona un'azienda valida.")

    return render(request, 'aziende/seleziona_azienda_timesheet.html', {'aziende': aziende})

def timesheet_per_azienda(request, id_azienda):
    # Recupera l'azienda specifica
    azienda = get_object_or_404(Azienda, pk=id_azienda)

    # Recupera i dipendenti associati all'azienda
    dipendenti = azienda.dipendenti.all()

    # Recupera tutti i timesheet associati ai dipendenti di questa azienda
    timesheets = Timesheet.objects.filter(dipendente__in=dipendenti).select_related('dipendente', 'piano_modulo')

    context = {
        'azienda': azienda,
        'timesheets': timesheets,
    }

    return render(request, 'aziende/timesheet_per_azienda.html', context)

def salva_timesheet(request, timesheet_id):
    if request.method == "POST":
        data = json.loads(request.body)
        timesheet = get_object_or_404(Timesheet, pk=timesheet_id)

        timesheet.ore_previste = data.get("ore_previste")
        timesheet.ore_effettuate = data.get("ore_effettuate")
        timesheet.importo_retributivo = data.get("importo_retributivo")
        timesheet.importo_contributivo = data.get("importo_contributivo")
        timesheet.save()

        return JsonResponse({"success": True, "message": "Timesheet aggiornato correttamente."})
    return JsonResponse({"success": False, "message": "Metodo non consentito."})


def salva_tutti_timesheets(request):
    if request.method == "POST":
        try:
            # Carica i dati inviati nella richiesta
            data = json.loads(request.body)
            updates = data.get("updates", [])
            print("Dati ricevuti:", updates)  # Stampa i dati nel log per verificare

            for update in updates:
                # Recupera il timesheet dal database
                timesheet = Timesheet.objects.get(pk=update["id"])
                
                # Aggiorna i campi del timesheet
                timesheet.ore_previste = update.get("ore_previste", timesheet.ore_previste)
                timesheet.ore_effettuate = update.get("ore_effettuate", timesheet.ore_effettuate)
                timesheet.importo_retributivo = update.get("importo_retributivo", timesheet.importo_retributivo)
                timesheet.importo_contributivo = update.get("importo_contributivo", timesheet.importo_contributivo)
                timesheet.save()

            return JsonResponse({"success": True, "message": "Tutti i timesheet sono stati aggiornati correttamente."})
        except Timesheet.DoesNotExist:
            return JsonResponse({"success": False, "message": "Un timesheet specificato non esiste."})
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Errore durante l'aggiornamento: {str(e)}"})
    return JsonResponse({"success": False, "message": "Metodo non consentito."})

def get_tipofondo_table(request):
    queryset = TipoFondo.objects.all().order_by('nome')

    fields = ['nome', 'percentuale_dad', 'percentuale_fad']
    manual_name_and_verbose = {
        'nome': ('nome', 'Nome'),
        'percentuale_dad': ('percentuale_dad', 'Percentuale DAD (%)'),
        'percentuale_fad': ('percentuale_fad', 'Percentuale FAD (%)'),
    }

    additional_fields = ['id']
    table_filter = 'global'
    paginate = sortable = True

    table = Table(
        request=request,
        queryset=queryset,
        fields=fields,
        additional_fields=additional_fields,
        manual_name_and_verbose=manual_name_and_verbose,
        table_filter=table_filter,
        sortable=sortable,
        paginate=paginate
    )

    return table

class TipoFondoListView(View):
    raise_exception = True

    def get(self, request, *args, **kwargs):
        context = {}
        template = 'tipofondo/tipofondo_list.html'

        table = get_tipofondo_table(request)

        if 'type' in request.GET and request.GET['type'] == 'view':
            return table.execute()

        context['table'] = [table.datatable_data_keys, table.table_columns_headers]
        context['field_header'] = zip(table.datatable_data_keys, table.table_columns_headers)

        return render(request, template, context)

class TipoFondoCreateView(CreateView):
    model = TipoFondo
    form_class = TipoFondoForm
    template_name = 'tipofondo/tipofondo_create_form.html'
    success_url = reverse_lazy('core:tipofondo_list')
    raise_exception = True

class TipoFondoUpdateView(UpdateView):
    model = TipoFondo
    form_class = TipoFondoForm
    template_name = 'tipofondo/tipofondo_edit_form.html'
    success_url = reverse_lazy('core:tipofondo_list')
    raise_exception = True

class TipoFondoDeleteView(DeleteView):
    model = TipoFondo
    template_name = 'tipofondo/tipofondo_confirm_delete.html'
    success_url = reverse_lazy('core:tipofondo_list')
    raise_exception = True

def search_tipofondo(request):
    search_term = request.GET.get('term', '')
    page = request.GET.get('page', 1)
    page_size = 10

    queryset = TipoFondo.objects.filter(nome__icontains=search_term).order_by('nome')
    total = queryset.count()
    start = (int(page) - 1) * page_size
    end = start + page_size

    results = [
        {'id': elem.pk, 'text': elem.nome} for elem in queryset[start:end]
    ]

    return JsonResponse({
        'results': results,
        'pagination': {'more': end < total}
    })

